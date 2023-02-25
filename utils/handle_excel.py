# -*- coding: utf-8 -*-
# @Time    : 2021/2/5 21:21
# @Author  : sunxuan
# @Site    : 
# @File    : handle_excel.py
import openpyxl
import os
from Config import config


class HandleExcel(object):
	def __init__(self):
		self.data_info = config.case_data_path
		###读取xlsx
		self.excel_obj = openpyxl.load_workbook(self.data_info)

	def get_sheet_data(self, sheet_index=None):
		'''
		:param sheet_index: sheet索引 默认为第一个sheet
		:return: 返回sheet对象
		'''
		##加载所有sheet页
		self.sheet_name = self.excel_obj.sheetnames
		if not sheet_index:
			sheet_index = 0
		sheet_data = self.excel_obj[self.sheet_name[sheet_index]]
		return sheet_data

	def get_cell_value(self, row, col):
		'''
		返回某个单元格内容
		:param rows: 行 从1 开始
		:param cols: 列 从1 开始
		:return:
		'''
		cell_data = self.get_sheet_data().cell(row=row, column=col).value
		return cell_data
	def get_row_values(self,row_index=None):
		'''
		返回某一行的所有数据
		:param row_index: 行 从1 开始
		:return:
		'''
		if not row_index:
			row_index=1
		return [row.value for row in self.get_sheet_data()[row_index]]
	def get_col_values(self,col_index=None):
		'''
		:param col_index: 列号 从'A'开始
		:return:  返回某一列所有数据
		'''
		if not col_index:
			col_index='A'
		return [col.value for col in self.get_sheet_data()[col_index]]

	def get_maxrows(self):
		'''
		:return: 返回最大行数
		'''
		row_count =self.get_sheet_data().max_row
		return  row_count
	def get_row_number(self,case_id):
		'''
		:param case_id: 根据case编号
		:return: 返回单元格所在的行数
		'''
		num=1
		cols_data =self.get_col_values()
		for col_data in cols_data:
			if col_data ==case_id:
				return  num
			num+=1
		return num



	def get_excel_data(self):
		'''
		:return:返回单元格所有数据 list
		'''
		data_list =[]
		for i in range(1,self.get_maxrows()):
			print(handle.get_row_values(i+1))
			data_list.append(handle.get_row_values(i+1))
		return  data_list

	def write_to_cell(self,row_index,col_index,value):
		'''

		:param row_index: 行
		:param col_index: 列
		:param value: 值
		:return: 写入数据
		'''
		wb =self.excel_obj
		wr =wb.active
		try:
			wr.cell(row_index, col_index).value = value
			wb.save(self.data_info)
		except:
			wr.cell(row_index, col_index).value = '写入失败'
			wb.save(self.data_info)

excel_data_instance =HandleExcel()

if __name__ == '__main__':
	handle = HandleExcel()
	print(handle.get_sheet_data())
	print('单元格内容：',handle.get_cell_value(2, 13))
	print('---->最大行数：',handle.get_maxrows())
	print('所有数据：',handle.get_excel_data())
	print('写入数据:',handle.write_to_cell(10,3,'write to data'))
	print(handle.get_row_values())
	print(handle.get_col_values())
	print('所在行数：',handle.get_row_number('imooc_007'))